import os

import openpyxl
import pandas as pd

file_path = 'USERS/users.xlsx'
df = pd.read_excel(file_path)
ids = df["Id"].tolist()

def create():
    # Создаем новую рабочую книгу (Excel-файл)
    workbook = openpyxl.Workbook()

    # Выбираем активный лист (первый лист по умолчанию)
    sheet = workbook.active

    # Записываем заголовки
    sheet['A1'] = 'Id'
    sheet['B1'] = 'Username'
    sheet['C1'] = 'Имя'
    sheet['D1'] = 'Фамилия'
    sheet['E1'] = 'Языковой код'


    # Сохраняем файл
    workbook.save('USERS/users.xlsx')



def clear():
    os.remove("USERS/users.xlsx")
    # Создаем новую рабочую книгу (Excel-файл)
    workbook = openpyxl.Workbook()

    # Выбираем активный лист (первый лист по умолчанию)
    sheet = workbook.active

    # Записываем заголовки
    sheet['A1'] = 'Id'
    sheet['B1'] = 'Username'
    sheet['C1'] = 'Имя'
    sheet['D1'] = 'Фамилия'
    sheet['E1'] = 'Языковой код'

    # Сохраняем файл
    workbook.save('USERS/users.xlsx')




def save_user(message):
    # Открываем существующий файл
    workbook = openpyxl.load_workbook('USERS/users.xlsx')

    # Выбираем активный лист (первый лист по умолчанию)
    sheet = workbook.active
    user = message.from_user
    # Пример данных для записи
    new_data = [
        (user.id, user.username, user.first_name, user.last_name, user.language_code)
    ]

    # Записываем новые данные в конец таблицы
    for row in new_data:
        sheet.append(row)

    # Сохраняем изменения в файле
    workbook.save('USERS\\users.xlsx')
